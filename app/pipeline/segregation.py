"""LLM-first file segregation engine (Step 30e-1).

Classifies files as PII vs non-PII using a vision LLM call on page 1-2.
Returns: pii yes/no, document type, field inventory, role attribution.

Two modes:
- Folder mode: classify every file, group by type, present for auditor review.
- Single-file mode: classify inline, skip grouping, proceed to analysis.

Design:
- One LLM call per file (~2-3 seconds).
- Vision model (qwen2.5vl:32b) for PDFs and images.
- Text fallback for non-renderable formats (XLSX, DOCX, MSG).
- Never raises — returns None on failure.
- Audit logged via OllamaClient.
"""
from __future__ import annotations

import base64
import io
import json
import logging
import os
import time
from dataclasses import dataclass, field, asdict
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# SegregationResult dataclass
# ---------------------------------------------------------------------------


@dataclass
class SegregationField:
    """A single PII field found during segregation."""
    name: str                          # field label as shown on document
    type: str                          # PERSON, US_SSN, LOCATION, etc.
    role: str = "primary_subject"      # primary_subject | secondary_contact
    value_visible: bool = True         # whether a value was visible

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class SegregationResult:
    """Result of LLM-first file classification."""
    file_path: str
    file_name: str
    file_type: str                     # pdf, xlsx, docx, msg, jpg, etc.
    total_pages: int

    # Classification
    pii_detected: bool = False
    confidence: float = 0.0
    document_type: str = "unknown"
    document_subtype: Optional[str] = None
    issuing_entity: Optional[str] = None
    primary_subject_type: Optional[str] = None
    summary: Optional[str] = None
    country_hint: Optional[str] = None  # ISO 3166-1 alpha-2 from LLM

    # Field inventory with role attribution
    fields: list[SegregationField] = field(default_factory=list)

    # Metadata
    llm_model_used: Optional[str] = None
    processing_time_ms: int = 0
    classification_method: str = "vision"  # vision | text | fallback
    error: Optional[str] = None
    raw_response: Optional[str] = None

    @property
    def field_inventory(self) -> list[str]:
        """List of field type strings (for grouping)."""
        return sorted(set(f.type for f in self.fields))

    @property
    def role_map(self) -> dict[str, str]:
        """Map field name → role (for downstream extraction)."""
        return {f.name: f.role for f in self.fields}

    @property
    def primary_fields(self) -> list[SegregationField]:
        """Fields belonging to the primary subject."""
        return [f for f in self.fields if f.role == "primary_subject"]

    @property
    def secondary_fields(self) -> list[SegregationField]:
        """Fields belonging to secondary contacts."""
        return [f for f in self.fields if f.role == "secondary_contact"]

    def to_dict(self) -> dict:
        d = asdict(self)
        d["field_inventory"] = self.field_inventory
        d["role_map"] = self.role_map
        return d


# ---------------------------------------------------------------------------
# SegregationEngine
# ---------------------------------------------------------------------------

# File types that can be rendered as images for vision LLM
_VISION_TYPES = frozenset({
    "pdf", "jpg", "jpeg", "png", "tiff", "tif", "bmp", "heic", "heif",
})

# File types with text extraction fallback
_TEXT_EXTRACTABLE_TYPES = frozenset({
    "xlsx", "xls", "xlsb", "csv", "tsv",
    "docx", "doc", "txt", "rtf",
    "msg", "eml",
})

# Max chars for text fallback prompt
_MAX_TEXT_CHARS = 4000


def _extract_pdf_text(file_path: str, max_pages: int = 2) -> str:
    """Extract text from the first N pages of a PDF using PyMuPDF.

    Returns concatenated text, or empty string on failure.
    Used to decide whether segregation can use text model (fast)
    instead of vision model (slow).
    """
    try:
        import fitz
        doc = fitz.open(file_path)
        parts = []
        for pg in range(min(max_pages, doc.page_count)):
            parts.append(doc[pg].get_text())
        doc.close()
        return "\n".join(parts)
    except Exception:
        return ""


class SegregationEngine:
    """Classifies files as PII vs non-PII using LLM.

    Usage:
        engine = SegregationEngine(db_session=session)
        result = engine.classify("path/to/file.pdf")
        if result and result.pii_detected:
            print(f"PII found: {result.field_inventory}")
    """

    def __init__(
        self,
        db_session=None,
        vision_model: Optional[str] = None,
        fallback_model: Optional[str] = None,
        project_id: Optional[str] = None,
    ):
        self._db_session = db_session
        self._vision_model = vision_model
        self._fallback_model = fallback_model
        self._project_id = project_id
        self._client = None  # lazy init
        self._corrections: list[dict] = []  # loaded correction memory
        if project_id:
            self._corrections = load_segregation_corrections(project_id)

    def _get_client(self):
        """Lazy-initialize OllamaClient."""
        if self._client is None:
            from app.llm.client import OllamaClient
            self._client = OllamaClient(db_session=self._db_session)
        return self._client

    def classify(
        self,
        file_path: str,
        document_id: Optional[str] = None,
    ) -> Optional[SegregationResult]:
        """Classify a single file. Returns SegregationResult or None on failure.

        Never raises exceptions to caller.
        """
        t0 = time.time()
        file_name = os.path.basename(file_path)
        file_type = _get_file_type(file_path)
        total_pages = _get_page_count(file_path, file_type)

        result = SegregationResult(
            file_path=file_path,
            file_name=file_name,
            file_type=file_type,
            total_pages=total_pages,
        )

        try:
            # For PDFs: check if text is extractable first. Text-based
            # classification with qwen2.5:7b is ~10x faster than vision
            # with 90B and equally accurate for text documents.  Only
            # fall back to vision when there's no extractable text.
            if file_type == "pdf":
                page_text = _extract_pdf_text(file_path, max_pages=2)
                if len(page_text.strip()) > 100:
                    self._classify_text(result, document_id, text_override=page_text)
                else:
                    self._classify_vision(result, document_id)
            elif file_type in _VISION_TYPES:
                self._classify_vision(result, document_id)
            elif file_type in _TEXT_EXTRACTABLE_TYPES:
                self._classify_text(result, document_id)
            else:
                # Unknown file type — mark as needing manual review
                result.error = f"Unsupported file type: {file_type}"
                result.classification_method = "fallback"
                logger.warning(
                    "Segregation: unsupported type %s for %s",
                    file_type, file_name,
                )
        except Exception:
            logger.exception("Segregation failed for %s", file_name)
            result.error = "Classification failed — see logs"
            result.classification_method = "fallback"

        # If LLM failed and result defaulted to non-PII, try regex fallback
        # so documents with obvious PII patterns aren't misclassified.
        if not result.pii_detected and result.classification_method == "fallback":
            text_for_fallback = None
            if file_type == "pdf":
                text_for_fallback = _extract_pdf_text(file_path, max_pages=2)
            elif file_type in _TEXT_EXTRACTABLE_TYPES:
                text_for_fallback = self._extract_text(file_path, file_type)
            if text_for_fallback:
                self._classify_regex_fallback(result, text_for_fallback)

        # Adaptive onset: if still non-PII but 50+ pages, sample mid-document
        # pages. Late-onset PII (e.g., K-1 schedules starting at page 52 in
        # a 100-page tax return) is invisible to pages 1-2 sampling.
        if (
            not result.pii_detected
            and total_pages >= 50
            and file_type == "pdf"
        ):
            self._check_late_onset_pii(result, file_path, total_pages, document_id)

        # Apply stored corrections from previous auditor reviews
        if self._corrections:
            result = apply_corrections(result, self._corrections)

        result.processing_time_ms = int((time.time() - t0) * 1000)
        return result

    def classify_batch(
        self,
        file_paths: list[str],
        progress_callback=None,
    ) -> list[SegregationResult]:
        """Classify multiple files. Returns list of results (never None entries)."""
        results = []
        for i, fp in enumerate(file_paths):
            result = self.classify(fp)
            if result:
                results.append(result)
            if progress_callback:
                progress_callback(i + 1, len(file_paths), result)
        return results

    # ----- Vision classification (PDFs, images) -----

    def _classify_vision(
        self,
        result: SegregationResult,
        document_id: Optional[str],
    ) -> None:
        """Classify using vision LLM on page 1 image."""
        from app.llm.prompts import SEGREGATION_PROMPT_VISION, SYSTEM_PROMPT

        image_b64 = self._render_page(result.file_path, result.file_type, page=0)
        if not image_b64:
            # Can't render — try text fallback
            logger.info(
                "Vision render failed for %s, trying text fallback",
                result.file_name,
            )
            self._classify_text(result, document_id)
            return

        prompt = SEGREGATION_PROMPT_VISION.format(
            file_name=result.file_name,
            file_type=result.file_type,
            total_pages=result.total_pages,
        )

        # Inject few-shot corrections from previous runs
        if self._corrections:
            prompt = _inject_corrections_into_prompt(prompt, self._corrections)

        client = self._get_client()

        # Try primary vision model
        response_text = None
        model_used = None
        try:
            response_text = client.generate_with_images(
                prompt=prompt,
                images=[image_b64],
                use_case="segregation",
                document_id=document_id,
                model_override=self._vision_model,
            )
            model_used = self._vision_model or "default_vision"
        except Exception as e:
            logger.warning(
                "Segregation primary vision failed for %s: %s",
                result.file_name, e,
            )

        # Fallback model if primary failed
        if not response_text and self._fallback_model:
            try:
                response_text = client.generate_with_images(
                    prompt=prompt,
                    images=[image_b64],
                    use_case="segregation_fallback",
                    document_id=document_id,
                    model_override=self._fallback_model,
                )
                model_used = self._fallback_model
            except Exception as e:
                logger.warning(
                    "Segregation fallback vision failed for %s: %s",
                    result.file_name, e,
                )

        if not response_text:
            result.error = "Vision LLM returned no response"
            result.classification_method = "fallback"
            return

        result.raw_response = response_text
        result.llm_model_used = model_used
        result.classification_method = "vision"
        self._parse_response(response_text, result)

        # If page 1 showed no PII but doc has >1 page, try page 2
        if (
            not result.pii_detected
            and result.total_pages > 1
            and result.confidence < 0.8
        ):
            self._retry_page2_vision(result, document_id, prompt)

    def _retry_page2_vision(
        self,
        result: SegregationResult,
        document_id: Optional[str],
        prompt: str,
    ) -> None:
        """Try page 2 if page 1 didn't show PII (e.g., cover page)."""
        image_b64 = self._render_page(result.file_path, result.file_type, page=1)
        if not image_b64:
            return

        client = self._get_client()
        try:
            response_text = client.generate_with_images(
                prompt=prompt,
                images=[image_b64],
                use_case="segregation_page2",
                document_id=document_id,
                model_override=self._vision_model or self._fallback_model,
            )
            if response_text:
                page2_result = SegregationResult(
                    file_path=result.file_path,
                    file_name=result.file_name,
                    file_type=result.file_type,
                    total_pages=result.total_pages,
                )
                self._parse_response(response_text, page2_result)
                # If page 2 found PII, use its result
                if page2_result.pii_detected:
                    result.pii_detected = True
                    result.confidence = page2_result.confidence
                    result.document_type = page2_result.document_type
                    result.document_subtype = page2_result.document_subtype
                    result.issuing_entity = page2_result.issuing_entity
                    result.primary_subject_type = page2_result.primary_subject_type
                    result.summary = page2_result.summary
                    result.fields = page2_result.fields
                    result.raw_response = response_text
                    logger.info(
                        "Page 2 found PII for %s (page 1 was clean)",
                        result.file_name,
                    )
        except Exception:
            logger.debug("Page 2 retry failed for %s", result.file_name)

    # ----- Text classification (XLSX, DOCX, MSG, etc.) -----

    def _classify_text(
        self,
        result: SegregationResult,
        document_id: Optional[str],
        text_override: str | None = None,
    ) -> None:
        """Classify using text extraction + text LLM."""
        from app.llm.prompts import SEGREGATION_PROMPT_TEXT, SYSTEM_PROMPT

        text = text_override or self._extract_text(result.file_path, result.file_type)
        if not text or len(text.strip()) < 20:
            result.error = "Could not extract text from file"
            result.classification_method = "fallback"
            return

        # Truncate to budget
        text_truncated = text[:_MAX_TEXT_CHARS]

        prompt = SEGREGATION_PROMPT_TEXT.format(
            file_name=result.file_name,
            file_type=result.file_type,
            total_pages=result.total_pages,
            char_count=len(text_truncated),
            document_text=text_truncated,
        )

        # Inject few-shot corrections from previous runs
        if self._corrections:
            prompt = _inject_corrections_into_prompt(prompt, self._corrections)

        client = self._get_client()
        response_text = None
        try:
            response_text = client.generate(
                prompt=prompt,
                system=None,  # uses default SYSTEM_PROMPT
                use_case="segregation_text",
                document_id=document_id,
            )
        except Exception as e:
            logger.warning(
                "Segregation text LLM failed for %s: %s",
                result.file_name, e,
            )

        if not response_text:
            result.error = "Text LLM returned no response"
            result.classification_method = "fallback"
            return

        result.raw_response = response_text
        result.llm_model_used = "text_model"
        result.classification_method = "text"
        self._parse_response(response_text, result)

    # ----- Regex fallback when LLM is unavailable -----

    def _classify_regex_fallback(self, result: SegregationResult, text: str) -> None:
        """Deterministic PII pattern scan when LLM is unavailable.

        Scans extracted text for common PII patterns (SSN, DOB, phone,
        email, driver's license). If enough patterns match, classifies
        as PII with moderate confidence so it doesn't silently default
        to non-PII when Ollama is down.
        """
        import re

        patterns = {
            "US_SSN": re.compile(r"\b\d{3}[-‐]?\d{2}[-‐]?\d{4}\b"),
            "GOV_ID": re.compile(
                r"\b(?:[A-Z]{2}\d{6}[A-Z]|\d{3}-\d{2}-\d{4}|[A-Z]\d{4,8})\b"
            ),
            "DATE_OF_BIRTH": re.compile(
                r"\b(?:DOB|Date of Birth|Birth\s*Date|BIRTHDATE)\b", re.IGNORECASE
            ),
            "PHONE_NUMBER": re.compile(
                r"\b(?:\(\d{3}\)\s*|\d{3}[-.])\d{3}[-.]?\d{4}\b"
            ),
            "EMAIL_ADDRESS": re.compile(
                r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
            ),
            "US_DRIVER_LICENSE": re.compile(
                r"\b(?:DL|Driver'?s?\s*Lic|License\s*#?)\b", re.IGNORECASE
            ),
            "PERSON": re.compile(
                r"\b(?:Name|Member(?:'s)?\s*Name|Employee\s*Name|Patient\s*Name"
                r"|Student\s*Name|(?:Mr|Mrs|Ms|Miss|Dr)\s+[A-Z])\b",
                re.IGNORECASE,
            ),
            "LOCATION": re.compile(
                r"\b(?:Address|Street|City|State|Zip|Postcode|Post\s*Code)\s*[:]\s*",
                re.IGNORECASE,
            ),
            "FINANCIAL": re.compile(
                r"\b(?:Account\s*(?:No|Number|#)|Sort\s*Code|IBAN|Transfer\s*Value"
                r"|Pension|Entitlement)\b",
                re.IGNORECASE,
            ),
        }

        found_types: list[str] = []
        for pii_type, pattern in patterns.items():
            if pattern.search(text):
                found_types.append(pii_type)

        if len(found_types) >= 2:
            result.pii_detected = True
            result.confidence = min(0.6 + len(found_types) * 0.05, 0.85)
            result.classification_method = "regex_fallback"
            result.document_type = "pii_document"
            result.error = (
                f"LLM unavailable — classified via regex fallback "
                f"({len(found_types)} PII patterns found: {', '.join(found_types)})"
            )
            result.fields = [
                SegregationField(name=t, type=t, role="primary_subject")
                for t in found_types
            ]
            logger.info(
                "Regex fallback: %s classified as PII (%d patterns: %s)",
                result.file_name, len(found_types), ", ".join(found_types),
            )
        elif len(found_types) == 1:
            # Single pattern — flag for review, don't auto-classify
            result.error = (
                f"LLM unavailable — 1 PII pattern found ({found_types[0]}), "
                f"needs manual review"
            )
            logger.info(
                "Regex fallback: %s has 1 PII pattern (%s) — needs review",
                result.file_name, found_types[0],
            )
        else:
            result.error = "LLM unavailable — no PII patterns detected by regex"
            logger.info(
                "Regex fallback: %s — no PII patterns found",
                result.file_name,
            )

    # ----- Adaptive onset: mid-document PII sampling -----

    def _check_late_onset_pii(
        self,
        result: SegregationResult,
        file_path: str,
        total_pages: int,
        document_id: Optional[str],
    ) -> None:
        """Sample mid-document pages for PII when pages 1-2 showed none.

        Some documents (tax returns, legal filings) have 50+ pages of
        boilerplate before individual PII begins. This binary-search-style
        sampling catches those cases without reading the entire document.
        """
        import re

        # Sample pages at 25%, 50%, 75% of the document
        sample_indices = sorted(set([
            total_pages // 4,
            total_pages // 2,
            (total_pages * 3) // 4,
        ]))

        try:
            import fitz
            doc = fitz.open(file_path)
        except Exception:
            return

        pii_page = None
        pii_text = None
        ssn_re = re.compile(r"\b\d{3}[-‐]?\d{2}[-‐]?\d{4}\b")
        name_re = re.compile(
            r"\b(?:Name|Member|Employee|Patient|Partner|Beneficiary)\s*[:]\s*",
            re.IGNORECASE,
        )

        try:
            for pg_idx in sample_indices:
                if pg_idx >= doc.page_count:
                    continue
                text = doc[pg_idx].get_text()
                doc._forget_page(pg_idx)
                if not text or len(text.strip()) < 30:
                    continue
                # Quick PII pattern check
                has_ssn = bool(ssn_re.search(text))
                has_name = bool(name_re.search(text))
                if has_ssn or has_name:
                    pii_page = pg_idx
                    pii_text = text
                    break
        finally:
            doc.close()

        if pii_page is None:
            return

        # Found PII in mid-document — try LLM classification on that page
        try:
            client = self._get_client()
            from app.llm.prompts import SEGREGATION_PROMPT_TEXT
            prompt = SEGREGATION_PROMPT_TEXT.format(
                file_name=result.file_name,
                file_type=result.file_type,
                total_pages=total_pages,
                char_count=len(pii_text[:_MAX_TEXT_CHARS]),
                document_text=pii_text[:_MAX_TEXT_CHARS],
            )
            response_text = client.generate(
                prompt=prompt,
                system=None,
                use_case="segregation_late_onset",
                document_id=document_id,
            )
            if response_text:
                self._parse_response(response_text, result)
                if result.pii_detected:
                    result.classification_method = "text_late_onset"
                    result.summary = (
                        f"Late-onset PII found at page {pii_page + 1} "
                        f"(pages 1-2 were non-PII boilerplate). "
                        f"{result.summary or ''}"
                    )
                    logger.info(
                        "Late onset: %s has PII starting at page %d (of %d)",
                        result.file_name, pii_page + 1, total_pages,
                    )
                    return
        except Exception as e:
            logger.warning("Late onset LLM failed for %s: %s", result.file_name, e)

        # LLM unavailable — fall back to regex on the mid-doc page
        self._classify_regex_fallback(result, pii_text)
        if result.pii_detected:
            result.summary = (
                f"Late-onset PII detected via regex at page {pii_page + 1}. "
                f"{result.summary or ''}"
            )
            logger.info(
                "Late onset (regex): %s has PII at page %d",
                result.file_name, pii_page + 1,
            )

    # ----- Response parsing -----

    def _parse_response(self, response_text: str, result: SegregationResult) -> None:
        """Parse LLM JSON response into SegregationResult fields."""
        try:
            # Strip markdown code fences if present
            text = response_text.strip()
            if text.startswith("```"):
                lines = text.split("\n")
                # Remove first and last lines (fences)
                lines = [
                    l for l in lines
                    if not l.strip().startswith("```")
                ]
                text = "\n".join(lines)

            data = json.loads(text)
        except (json.JSONDecodeError, ValueError) as e:
            logger.warning(
                "Segregation: invalid JSON from LLM for %s: %s",
                result.file_name, e,
            )
            result.error = f"Invalid JSON response: {e}"
            return

        # Core fields
        result.pii_detected = bool(data.get("pii", False))
        result.confidence = float(data.get("confidence", 0.0))
        result.document_type = str(data.get("document_type", "unknown"))
        result.document_subtype = data.get("document_subtype")
        result.issuing_entity = data.get("issuing_entity")
        result.primary_subject_type = data.get("primary_subject_type")
        result.summary = data.get("summary")

        # Country hint (ISO 3166-1 alpha-2) — used downstream by the
        # government-ID classifier to disambiguate digit-only formats.
        ch = data.get("country_hint")
        if isinstance(ch, str) and len(ch.strip()) == 2:
            result.country_hint = ch.strip().upper()

        # Field inventory with roles
        raw_fields = data.get("fields", [])
        if isinstance(raw_fields, list):
            for f in raw_fields:
                if isinstance(f, dict):
                    result.fields.append(SegregationField(
                        name=str(f.get("name", "unknown")),
                        type=str(f.get("type", "OTHER_ID")),
                        role=str(f.get("role", "primary_subject")),
                        value_visible=bool(f.get("value_visible", True)),
                    ))

    # ----- Rendering helpers -----

    def _render_page(
        self,
        file_path: str,
        file_type: str,
        page: int = 0,
    ) -> Optional[str]:
        """Render a file page as base64 PNG. Returns None on failure."""
        try:
            if file_type == "pdf":
                from app.pdf.renderer import render_page_to_image
                return render_page_to_image(file_path, page, dpi=150)

            elif file_type in ("heic", "heif"):
                try:
                    from pillow_heif import register_heif_opener
                    register_heif_opener()
                except ImportError:
                    pass
                return self._pil_to_base64(file_path)

            elif file_type in ("jpg", "jpeg", "png", "tiff", "tif", "bmp"):
                return self._pil_to_base64(file_path)

        except Exception:
            logger.debug("Render failed for %s page %d", file_path, page)
        return None

    @staticmethod
    def _pil_to_base64(file_path: str) -> str:
        """Open an image file with PIL and return base64 PNG."""
        from PIL import Image
        img = Image.open(file_path).convert("RGB")
        # Resize if too large (keep manageable for LLM)
        max_dim = 2048
        if max(img.size) > max_dim:
            ratio = max_dim / max(img.size)
            new_size = (int(img.width * ratio), int(img.height * ratio))
            img = img.resize(new_size, Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("utf-8")

    # ----- Text extraction helpers -----

    def _extract_text(self, file_path: str, file_type: str) -> Optional[str]:
        """Extract text from non-renderable file types."""
        try:
            if file_type in ("xlsx", "xls", "xlsb"):
                return self._extract_xlsx_text(file_path)
            elif file_type == "csv":
                return self._extract_csv_text(file_path)
            elif file_type == "docx":
                return self._extract_docx_text(file_path)
            elif file_type in ("txt", "rtf"):
                with open(file_path, "r", errors="replace") as f:
                    return f.read(_MAX_TEXT_CHARS)
            elif file_type == "msg":
                return self._extract_msg_text(file_path)
            elif file_type == "eml":
                return self._extract_eml_text(file_path)
        except Exception:
            logger.debug("Text extraction failed for %s", file_path)
        return None

    @staticmethod
    def _extract_xlsx_text(file_path: str) -> Optional[str]:
        """Extract text from first sheet of Excel file."""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            ws = wb.active
            lines = []
            for row in ws.iter_rows(max_row=50, values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                lines.append("\t".join(cells))
            wb.close()
            return "\n".join(lines)
        except Exception:
            pass
        # Fallback for xls/xlsb
        try:
            import xlrd
            wb = xlrd.open_workbook(file_path)
            ws = wb.sheet_by_index(0)
            lines = []
            for row_idx in range(min(50, ws.nrows)):
                cells = [str(ws.cell_value(row_idx, c)) for c in range(ws.ncols)]
                lines.append("\t".join(cells))
            return "\n".join(lines)
        except Exception:
            pass
        return None

    @staticmethod
    def _extract_csv_text(file_path: str) -> Optional[str]:
        """Extract text from CSV file."""
        with open(file_path, "r", errors="replace") as f:
            return f.read(_MAX_TEXT_CHARS)

    @staticmethod
    def _extract_docx_text(file_path: str) -> Optional[str]:
        """Extract text from Word document."""
        try:
            from docx import Document
            doc = Document(file_path)
            text = "\n".join(p.text for p in doc.paragraphs[:100])
            return text[:_MAX_TEXT_CHARS]
        except Exception:
            return None

    @staticmethod
    def _extract_msg_text(file_path: str) -> Optional[str]:
        """Extract text from Outlook MSG file."""
        try:
            import extract_msg
            msg = extract_msg.Message(file_path)
            parts = []
            if msg.subject:
                parts.append(f"Subject: {msg.subject}")
            if msg.sender:
                parts.append(f"From: {msg.sender}")
            if msg.body:
                parts.append(msg.body[:_MAX_TEXT_CHARS])
            msg.close()
            return "\n".join(parts)
        except Exception:
            return None

    @staticmethod
    def _extract_eml_text(file_path: str) -> Optional[str]:
        """Extract text from EML file."""
        import email
        with open(file_path, "r", errors="replace") as f:
            msg = email.message_from_file(f)
        parts = []
        parts.append(f"Subject: {msg.get('Subject', '')}")
        parts.append(f"From: {msg.get('From', '')}")
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    payload = part.get_payload(decode=True)
                    if payload:
                        parts.append(payload.decode("utf-8", errors="replace"))
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                parts.append(payload.decode("utf-8", errors="replace"))
        return "\n".join(parts)[:_MAX_TEXT_CHARS]


# ---------------------------------------------------------------------------
# Utility functions
# ---------------------------------------------------------------------------


def _get_file_type(file_path: str) -> str:
    """Extract file extension (lowercase, no dot)."""
    _, ext = os.path.splitext(file_path)
    return ext.lower().lstrip(".")


# ---------------------------------------------------------------------------
# Correction memory (Step 30e-5)
# ---------------------------------------------------------------------------


def load_segregation_corrections(project_id: str) -> list[dict]:
    """Load correction memory from the project's JSONL file.

    Returns a list of correction dicts with keys:
    group_id, group_name, document_type, action (reject/reclassify),
    new_document_type, new_is_pii, corrected_at, etc.
    """
    try:
        from app.core.settings import get_settings
        from pathlib import Path
        settings = get_settings()
        corrections_dir = Path(settings.upload_dir).parent / "corrections"
        filepath = corrections_dir / f"{project_id}_segregation_corrections.jsonl"
        if not filepath.exists():
            return []
        corrections = []
        with open(filepath) as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        corrections.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        logger.info(
            "Loaded %d segregation corrections for project %s",
            len(corrections), project_id,
        )
        return corrections
    except Exception as e:
        logger.warning("Failed to load segregation corrections: %s", e)
        return []


def apply_corrections(
    result: SegregationResult,
    corrections: list[dict],
) -> SegregationResult:
    """Apply stored corrections to a new segregation result.

    Checks if a previous correction matches this file's document_type.
    If so, overrides the LLM's classification with the human correction.

    This handles:
    - Documents that were reclassified (wrong type → correct type)
    - Documents that were rescued from non-PII (false negatives)
    - Documents that were rejected (false positives)
    """
    if not corrections:
        return result

    for corr in corrections:
        action = corr.get("action")
        old_doc_type = corr.get("document_type")

        # Match by document type (the most reliable signal)
        if old_doc_type and old_doc_type == result.document_type:
            if action == "reclassify":
                new_type = corr.get("new_document_type")
                new_is_pii = corr.get("new_is_pii")
                if new_type:
                    logger.info(
                        "Applying correction: %s → %s for %s",
                        result.document_type, new_type, result.file_name,
                    )
                    result.document_type = new_type
                if new_is_pii is not None:
                    result.pii_detected = bool(new_is_pii)
                return result  # apply first matching correction

            elif action == "reject":
                # This type was rejected before — mark as non-PII
                logger.info(
                    "Applying rejection for %s type: %s",
                    result.file_name, old_doc_type,
                )
                result.pii_detected = False
                return result

    return result


def _inject_corrections_into_prompt(
    prompt: str,
    corrections: list[dict],
    max_examples: int = 5,
) -> str:
    """Inject few-shot correction examples into a segregation prompt.

    Adds a section like:
    "IMPORTANT: In previous runs on similar documents, the following
    corrections were made by the auditor. Please consider them:
    - Documents of type 'shipping_document' were marked as NOT containing PII.
    - Documents of type 'medical_intake' were reclassified from 'invoice'.
    "
    """
    if not corrections:
        return prompt

    # Deduplicate corrections by (old_type, action)
    seen: set[tuple[str, str]] = set()
    unique_corrections: list[dict] = []
    for c in reversed(corrections):  # most recent first
        key = (c.get("document_type", ""), c.get("action", ""))
        if key not in seen:
            seen.add(key)
            unique_corrections.append(c)
            if len(unique_corrections) >= max_examples:
                break

    if not unique_corrections:
        return prompt

    lines = [
        "\n\nIMPORTANT: In previous auditor reviews of similar documents, "
        "the following corrections were made. Apply these lessons:\n"
    ]
    for c in unique_corrections:
        action = c.get("action")
        doc_type = c.get("document_type", "unknown")
        if action == "reject":
            lines.append(
                f"- Documents of type '{doc_type}' do NOT contain PII "
                f"(previously incorrectly classified as PII)."
            )
        elif action == "reclassify":
            new_type = c.get("new_document_type", "unknown")
            new_pii = c.get("new_is_pii")
            pii_note = ""
            if new_pii is not None:
                pii_note = f" (PII: {'yes' if new_pii else 'no'})"
            lines.append(
                f"- Documents of type '{doc_type}' should be classified "
                f"as '{new_type}'{pii_note}."
            )
    lines.append("")  # trailing newline

    # Insert before the JSON response format
    injection = "\n".join(lines)
    # Insert before "Respond with ONLY this JSON"
    marker = "Respond with ONLY this JSON"
    if marker in prompt:
        idx = prompt.index(marker)
        return prompt[:idx] + injection + prompt[idx:]
    else:
        return prompt + injection


def _get_page_count(file_path: str, file_type: str) -> int:
    """Get page count for supported file types."""
    if file_type == "pdf":
        try:
            import fitz
            doc = fitz.open(file_path)
            count = doc.page_count
            doc.close()
            return count
        except Exception:
            return 0
    elif file_type in ("jpg", "jpeg", "png", "bmp", "heic", "heif", "tiff", "tif"):
        return 1
    elif file_type in ("xlsx", "xls", "xlsb"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True)
            count = len(wb.sheetnames)
            wb.close()
            return count
        except Exception:
            return 1
    return 1
