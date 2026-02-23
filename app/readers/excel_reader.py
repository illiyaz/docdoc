"""Excel reader: openpyxl multi-tab streaming.

Rules (CLAUDE.md § 2 — Excel Multi-Tab Reader)
-----------------------------------------------
- Always open with read_only=True, data_only=True — never eager load.
- Each worksheet is an independent context window; PageStitcher.reset()
  is called before every visible sheet so context never bleeds across tabs.
- Row 1 of each sheet is the header row; values become col_header on every
  block in that column and are emitted as block_type="table_header".
- Rows 2+ are data cells emitted as block_type="table_cell".
- Hidden or empty sheets are skipped — no blocks are created for them.
- Tab name is stored as page_or_sheet (str) on every ExtractedBlock.
- bbox=None for all blocks (non-visual format).

False-positive guard
--------------------
If >80% of non-empty values in a column match the same structured-ID
regex pattern (e.g. SSN XXX-XX-XXXX, product SKU, long numeric IDs),
the column is flagged as a structured ID field rather than PII.
Flagged cells keep block_type="table_cell" but their col_header is
prefixed with "[REVIEW] " to signal downstream handling.
"""
from __future__ import annotations

import re
import uuid
from pathlib import Path

import openpyxl

from app.readers.base import BaseReader, ExtractedBlock
from app.readers.stitcher import PageStitcher

# Regex patterns whose prevalence (>80%) in a column indicates structured IDs,
# not organic PII.  Each is a full-match pattern (anchored by re.fullmatch).
_STRUCTURED_ID_PATTERNS: list[str] = [
    r"\d{3}-\d{2}-\d{4}",          # SSN-style:          123-45-6789
    r"\d{4}-\d{4}-\d{4}-\d{4}",    # credit-card-style:  1234-5678-9012-3456
    r"\d{3}-\d{3}-\d{4}",          # phone-style:        123-456-7890
    r"[A-Z]{2,3}-\d{4,}",          # product-ID-style:   AB-12345
    r"\d{6,}",                      # long numeric row IDs: 123456789
]


class ExcelReader(BaseReader):
    """Stream an Excel workbook sheet-by-sheet and emit ExtractedBlock objects."""

    def __init__(self, path: str | Path) -> None:
        super().__init__(path)

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def read(self) -> list[ExtractedBlock]:
        """Iterate over all visible, non-empty sheets and return blocks.

        Sheets are processed in workbook order.  Hidden and veryHidden sheets
        are skipped without touching the stitcher.  PageStitcher.reset() is
        called before each visible sheet to prevent cross-tab context bleed.
        """
        stitcher = PageStitcher()
        all_blocks: list[ExtractedBlock] = []

        wb = openpyxl.load_workbook(str(self.path), read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                if ws.sheet_state != "visible":
                    continue  # skip hidden / veryHidden — no reset, no blocks

                stitcher.reset()  # isolate each visible sheet's context
                sheet_blocks = self._read_sheet(ws, sheet_name)
                all_blocks.extend(sheet_blocks)
        finally:
            wb.close()

        return all_blocks

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _read_sheet(self, sheet: object, sheet_name: str) -> list[ExtractedBlock]:
        """Extract ExtractedBlock objects from a single worksheet.

        Returns an empty list for sheets that have no rows or only None values
        (treated as empty — CLAUDE.md Excel rule: empty sheets are skipped).
        """
        source = str(self.path)
        file_type = self.path.suffix.lstrip(".").lower() or "xlsx"

        rows = list(sheet.iter_rows())
        if not rows:
            return []

        # A sheet where every cell in every row is None is considered empty.
        if all(cell.value is None for row in rows for cell in row):
            return []

        table_id = str(uuid.uuid4())

        # ---- Row 1: column headers ----------------------------------------
        header_row = rows[0]
        # Map 1-based column index → header text for use on data cells
        headers: dict[int, str] = {}
        header_blocks: list[ExtractedBlock] = []
        for cell in header_row:
            header_text = str(cell.value) if cell.value is not None else ""
            headers[cell.column] = header_text
            header_blocks.append(ExtractedBlock(
                text=header_text,
                page_or_sheet=sheet_name,
                source_path=source,
                file_type=file_type,
                block_type="table_header",
                bbox=None,
                row=cell.row,
                column=cell.column,
                table_id=table_id,
                col_header=header_text,
                row_index=0,
            ))

        # ---- Rows 2+: data cells ------------------------------------------
        data_blocks: list[ExtractedBlock] = []
        for row in rows[1:]:
            for cell in row:
                text = str(cell.value) if cell.value is not None else ""
                col_header = headers.get(cell.column, "")
                data_blocks.append(ExtractedBlock(
                    text=text,
                    page_or_sheet=sheet_name,
                    source_path=source,
                    file_type=file_type,
                    block_type="table_cell",
                    bbox=None,
                    row=cell.row,
                    column=cell.column,
                    table_id=table_id,
                    col_header=col_header,
                    row_index=cell.row - 1,  # 0-based: header is 0, first data row is 1
                ))

        # ---- False-positive guard -----------------------------------------
        # Group data cell texts by column, then check each column against
        # every structured-ID pattern.  Flag the column if >80% match.
        col_values: dict[int, list[str]] = {}
        for block in data_blocks:
            col_values.setdefault(block.column, []).append(block.text)

        flagged_cols: set[int] = set()
        for col_idx, values in col_values.items():
            for pattern in _STRUCTURED_ID_PATTERNS:
                if self._is_structured_id_column(values, pattern):
                    flagged_cols.add(col_idx)
                    break  # one matching pattern is enough to flag the column

        for block in data_blocks:
            if block.column in flagged_cols:
                block.col_header = f"[REVIEW] {block.col_header}"

        return header_blocks + data_blocks

    def _is_structured_id_column(self, column_values: list[str], pattern: str) -> bool:
        """Return True if strictly more than 80% of non-empty values match pattern.

        Parameters
        ----------
        column_values:
            All text values in the column (may include empty strings).
        pattern:
            A regex pattern used with re.fullmatch (full-string match,
            case-insensitive).
        """
        non_empty = [v for v in column_values if v.strip()]
        if not non_empty:
            return False
        matching = sum(
            1 for v in non_empty if re.fullmatch(pattern, v.strip(), re.IGNORECASE)
        )
        return matching / len(non_empty) > 0.80
